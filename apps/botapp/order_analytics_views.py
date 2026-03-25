"""
Order Analytics Admin Views.
Separate admin panel section for order tracking and worker performance.
"""
import json
import asyncio
import io
from datetime import datetime, timedelta
from decimal import Decimal
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages

# Import Django models
from apps.botapp.models import WorkerServiceFee

# Import analytics functions
from bot.services.order_analytics_tracker import (
    get_order_analytics_summary,
    get_top_workers,
    get_fastest_workers,
    get_revenue_stats,
    get_daily_orders_chart,
    get_daily_revenue_chart,
    get_orders_list,
    get_order_details,
    get_worker_details,
    get_cemetery_stats,
    get_worker_earnings_by_period,
    get_period_orders_detail,
    get_period_summary,
)


def run_async(coro):
    """Run async function in sync context."""
    try:
        loop = asyncio.get_event_loop()
    except RuntimeError:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)

    if loop.is_running():
        import concurrent.futures
        with concurrent.futures.ThreadPoolExecutor() as executor:
            future = executor.submit(asyncio.run, coro)
            return future.result()
    else:
        return loop.run_until_complete(coro)


@staff_member_required
def order_analytics_dashboard(request):
    """Main order analytics dashboard."""
    period = request.GET.get('period', '30d')

    if period == '7d':
        days = 7
        period_label = "7 kun"
    elif period == '90d':
        days = 90
        period_label = "90 kun"
    else:
        days = 30
        period_label = "30 kun"

    try:
        # Get analytics data
        summary = run_async(get_order_analytics_summary(days))
        top_workers = run_async(get_top_workers(10))
        fastest_workers = run_async(get_fastest_workers(10))
        revenue_stats = run_async(get_revenue_stats(days))
        daily_orders = run_async(get_daily_orders_chart(days))
        daily_revenue = run_async(get_daily_revenue_chart(days))
        cemetery_stats = run_async(get_cemetery_stats(days))

    except Exception as e:
        summary = {
            'total_orders': 0,
            'completed_orders': 0,
            'completion_rate': 0,
            'total_revenue': 0,
            'avg_accept_time_formatted': '—',
            'fastest_accept_formatted': '—',
            'slowest_accept_formatted': '—',
            'active_workers': 0,
        }
        top_workers = []
        fastest_workers = []
        revenue_stats = {'daily': 0, 'weekly': 0, 'monthly': 0}
        daily_orders = []
        daily_revenue = []
        cemetery_stats = []

    # Prepare chart data
    orders_chart_data = json.dumps({
        'labels': [d['date'][5:] for d in daily_orders],  # MM-DD format
        'total': [d['total'] for d in daily_orders],
        'completed': [d['completed'] for d in daily_orders],
    })

    revenue_chart_data = json.dumps({
        'labels': [d['date'][5:] for d in daily_revenue],
        'revenue': [d['revenue'] for d in daily_revenue],
    })

    context = {
        'title': 'Buyurtmalar Analitikasi',
        'period': period,
        'period_label': period_label,

        # Summary stats
        'summary': summary,

        # Revenue
        'revenue': revenue_stats,

        # Workers
        'top_workers': top_workers,
        'fastest_workers': fastest_workers,

        # Cemetery stats
        'cemetery_stats': cemetery_stats[:10],

        # Chart data
        'orders_chart_data': orders_chart_data,
        'revenue_chart_data': revenue_chart_data,
    }

    return render(request, 'admin/botapp/order_analytics.html', context)


@staff_member_required
def order_analytics_orders(request):
    """Orders list with filters."""
    page = int(request.GET.get('page', 1))
    per_page = 30
    offset = (page - 1) * per_page

    search = request.GET.get('search', '').strip()
    status = request.GET.get('status', '')
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')

    # Parse dates
    date_from = None
    date_to = None
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        except ValueError:
            pass
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            pass

    try:
        orders, total_count = run_async(get_orders_list(
            limit=per_page,
            offset=offset,
            search=search if search else None,
            status=status if status else None,
            date_from=date_from,
            date_to=date_to,
        ))
    except Exception as e:
        orders = []
        total_count = 0

    # Pagination
    total_pages = (total_count + per_page - 1) // per_page
    has_next = page < total_pages
    has_prev = page > 1

    context = {
        'title': 'Buyurtmalar Ro\'yxati',
        'orders': orders,
        'search': search,
        'status': status,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'total_count': total_count,
        'page': page,
        'total_pages': total_pages,
        'has_next': has_next,
        'has_prev': has_prev,
        'page_range': range(max(1, page - 2), min(total_pages + 1, page + 3)),
        'statuses': ['new', 'posted', 'in_progress', 'completed', 'cancelled'],
    }

    return render(request, 'admin/botapp/order_analytics_orders.html', context)


@staff_member_required
def order_analytics_order_detail(request, order_id):
    """Detailed order view."""
    try:
        order = run_async(get_order_details(order_id))
    except Exception:
        order = None

    if not order:
        return render(request, 'admin/botapp/order_analytics_order_detail.html', {
            'title': 'Buyurtma topilmadi',
            'order': None,
        })

    context = {
        'title': f'Buyurtma #{order["original_id"]}',
        'order': order,
    }

    return render(request, 'admin/botapp/order_analytics_order_detail.html', context)


@staff_member_required
def order_analytics_workers(request):
    """Workers list and performance."""
    try:
        top_workers = run_async(get_top_workers(50))
        fastest_workers = run_async(get_fastest_workers(50))
    except Exception:
        top_workers = []
        fastest_workers = []

    # Merge lists and sort by total completed
    workers_dict = {}
    for w in top_workers:
        workers_dict[w['telegram_id']] = w
    for w in fastest_workers:
        if w['telegram_id'] not in workers_dict:
            workers_dict[w['telegram_id']] = w

    workers = sorted(
        workers_dict.values(),
        key=lambda x: x.get('total_completed', 0),
        reverse=True
    )

    context = {
        'title': 'Ishchilar Statistikasi',
        'workers': workers,
        'fastest_workers': fastest_workers[:10],
    }

    return render(request, 'admin/botapp/order_analytics_workers.html', context)


@staff_member_required
def order_analytics_worker_detail(request, telegram_id):
    """Detailed worker view."""
    try:
        worker = run_async(get_worker_details(telegram_id))
    except Exception:
        worker = None

    if not worker:
        return render(request, 'admin/botapp/order_analytics_worker_detail.html', {
            'title': 'Ishchi topilmadi',
            'worker': None,
        })

    context = {
        'title': f'Ishchi: {worker.get("display_name") or worker.get("username") or telegram_id}',
        'worker': worker,
    }

    return render(request, 'admin/botapp/order_analytics_worker_detail.html', context)


@staff_member_required
def order_analytics_api(request):
    """API endpoint for real-time data."""
    period = request.GET.get('period', '30d')

    if period == '7d':
        days = 7
    elif period == '90d':
        days = 90
    else:
        days = 30

    try:
        data = {
            'summary': run_async(get_order_analytics_summary(days)),
            'revenue': run_async(get_revenue_stats(days)),
            'timestamp': datetime.now().isoformat(),
        }
    except Exception as e:
        data = {'error': str(e)}

    return JsonResponse(data)


@staff_member_required
def order_analytics_salary(request):
    """Salary report page - calculate worker earnings for a period."""
    # Handle service fee form submission
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_service_fee':
            service_name = request.POST.get('service_name', '').strip()
            fee_amount = request.POST.get('fee_amount', '0')
            if service_name and fee_amount:
                try:
                    WorkerServiceFee.objects.create(
                        service_name=service_name,
                        fee_amount=Decimal(fee_amount),
                        is_active=True
                    )
                    messages.success(request, f"'{service_name}' hizmat to'lovi qo'shildi")
                except Exception as e:
                    messages.error(request, f"Xatolik: {e}")
            return redirect(request.get_full_path())

        elif action == 'delete_service_fee':
            fee_id = request.POST.get('fee_id')
            if fee_id:
                try:
                    WorkerServiceFee.objects.filter(id=fee_id).delete()
                    messages.success(request, "Hizmat to'lovi o'chirildi")
                except Exception as e:
                    messages.error(request, f"Xatolik: {e}")
            return redirect(request.get_full_path())

        elif action == 'toggle_service_fee':
            fee_id = request.POST.get('fee_id')
            if fee_id:
                try:
                    fee = WorkerServiceFee.objects.get(id=fee_id)
                    fee.is_active = not fee.is_active
                    fee.save()
                except Exception as e:
                    pass
            return redirect(request.get_full_path())

    # Get period type and dates
    period_type = request.GET.get('period_type', 'week')
    commission_percent = float(request.GET.get('commission', 0))

    # Calculate default dates
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    if period_type == 'month':
        # Start of current month
        default_from = today.replace(day=1)
        # End of current month
        if today.month == 12:
            default_to = today.replace(year=today.year + 1, month=1, day=1)
        else:
            default_to = today.replace(month=today.month + 1, day=1)
        period_label = "Oylik hisobot"
    else:
        # Start of current week (Monday)
        default_from = today - timedelta(days=today.weekday())
        default_to = default_from + timedelta(days=7)
        period_label = "Haftalik hisobot"

    # Parse custom dates if provided
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        except ValueError:
            date_from = default_from
    else:
        date_from = default_from
        date_from_str = date_from.strftime('%Y-%m-%d')

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            date_to = default_to
    else:
        date_to = default_to
        date_to_str = (date_to - timedelta(days=1)).strftime('%Y-%m-%d')

    # Get service fees
    service_fees = list(WorkerServiceFee.objects.all().order_by('service_name'))
    active_service_fees = [sf for sf in service_fees if sf.is_active]

    # Calculate total service fee per order
    total_service_fee_per_order = sum(sf.fee_amount for sf in active_service_fees)

    # Get data
    try:
        workers = run_async(get_worker_earnings_by_period(date_from, date_to))
        summary = run_async(get_period_summary(date_from, date_to))
    except Exception as e:
        workers = []
        summary = {'total_orders': 0, 'total_revenue': 0, 'active_workers': 0}

    # Calculate commission and service fees for each worker
    for w in workers:
        w['commission_amount'] = round(w['total_earnings'] * commission_percent / 100, 2)
        # Service fee = per order fee * number of orders completed
        w['service_fee_total'] = round(float(total_service_fee_per_order) * w['order_count'], 2)
        # Net = earnings - commission + service fees
        w['net_payment'] = round(
            w['total_earnings'] - w['commission_amount'] + w['service_fee_total'],
            2
        )

    # Calculate totals
    total_earnings = sum(w['total_earnings'] for w in workers)
    total_commission = sum(w['commission_amount'] for w in workers)
    total_service_fees = sum(w['service_fee_total'] for w in workers)
    total_net = sum(w['net_payment'] for w in workers)

    context = {
        'title': 'Oylik/Haftalik Hisobot',
        'period_type': period_type,
        'period_label': period_label,
        'date_from': date_from_str,
        'date_to': date_to_str,
        'date_from_display': date_from.strftime('%d.%m.%Y'),
        'date_to_display': (date_to - timedelta(days=1)).strftime('%d.%m.%Y'),
        'commission_percent': commission_percent,
        'workers': workers,
        'summary': summary,
        'total_earnings': total_earnings,
        'total_commission': total_commission,
        'total_service_fees': total_service_fees,
        'total_net': total_net,
        # Service fees
        'service_fees': service_fees,
        'active_service_fees': active_service_fees,
        'total_service_fee_per_order': float(total_service_fee_per_order),
    }

    return render(request, 'admin/botapp/order_analytics_salary.html', context)


@staff_member_required
def order_analytics_salary_export(request):
    """Export salary report to Excel."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl", status=500)

    # Get parameters
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    commission_percent = float(request.GET.get('commission', 0))

    # Parse dates
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    default_from = today - timedelta(days=today.weekday())
    default_to = default_from + timedelta(days=7)

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        except ValueError:
            date_from = default_from
    else:
        date_from = default_from

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            date_to = default_to
    else:
        date_to = default_to

    # Get service fees
    active_service_fees = list(WorkerServiceFee.objects.filter(is_active=True))
    total_service_fee_per_order = sum(sf.fee_amount for sf in active_service_fees)

    # Get data
    try:
        workers = run_async(get_worker_earnings_by_period(date_from, date_to))
        orders = run_async(get_period_orders_detail(date_from, date_to))
        summary = run_async(get_period_summary(date_from, date_to))
    except Exception as e:
        return HttpResponse(f"Xatolik: {e}", status=500)

    # Create workbook
    wb = Workbook()

    # ===== Sheet 1: Summary =====
    ws1 = wb.active
    ws1.title = "Umumiy hisobot"

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
    service_fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
    title_font = Font(bold=True, size=14)
    money_font = Font(bold=True, color="059669")
    service_font = Font(bold=True, color="8b5cf6")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws1['A1'] = f"HAFTALIK/OYLIK HISOBOT"
    ws1['A1'].font = title_font
    ws1['A2'] = f"Davr: {date_from.strftime('%d.%m.%Y')} - {(date_to - timedelta(days=1)).strftime('%d.%m.%Y')}"
    ws1['A3'] = f"Komissiya: {commission_percent}%"
    ws1['A4'] = f"Yaratilgan: {datetime.now().strftime('%d.%m.%Y %H:%M')}"

    # Summary stats
    ws1['A6'] = "UMUMIY STATISTIKA"
    ws1['A6'].font = title_font
    ws1['A7'] = "Jami buyurtmalar:"
    ws1['B7'] = summary['total_orders']
    ws1['A8'] = "Jami daromad:"
    ws1['B8'] = f"{summary['total_revenue']:,.0f} so'm"
    ws1['A9'] = "Faol ishchilar:"
    ws1['B9'] = summary['active_workers']

    # Service fees section
    current_row = 11
    if active_service_fees:
        ws1.cell(row=current_row, column=1, value="HIZMAT TO'LOVLARI (har bir buyurtma uchun)")
        ws1.cell(row=current_row, column=1).font = title_font
        current_row += 1

        for sf in active_service_fees:
            ws1.cell(row=current_row, column=1, value=sf.service_name)
            ws1.cell(row=current_row, column=2, value=f"{sf.fee_amount:,.0f} so'm")
            ws1.cell(row=current_row, column=2).font = service_font
            current_row += 1

        ws1.cell(row=current_row, column=1, value="JAMI (buyurtma uchun):")
        ws1.cell(row=current_row, column=1).font = Font(bold=True)
        ws1.cell(row=current_row, column=2, value=f"{total_service_fee_per_order:,.0f} so'm")
        ws1.cell(row=current_row, column=2).font = Font(bold=True, color="8b5cf6")
        current_row += 2

    # Workers table header
    ws1.cell(row=current_row, column=1, value="ISHCHILAR HISOBOTI")
    ws1.cell(row=current_row, column=1).font = title_font
    current_row += 1

    headers = ['#', 'Ishchi', 'Username', 'Telegram ID', 'Buyurtmalar', 'Jami summa',
               'Komissiya', 'Hizmat to\'lovi', 'To\'lov summa']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    header_row = current_row
    current_row += 1

    # Workers data
    for idx, w in enumerate(workers, 1):
        w['commission_amount'] = round(w['total_earnings'] * commission_percent / 100, 2)
        w['service_fee_total'] = round(float(total_service_fee_per_order) * w['order_count'], 2)
        w['net_payment'] = round(w['total_earnings'] - w['commission_amount'] + w['service_fee_total'], 2)

        row = current_row
        ws1.cell(row=row, column=1, value=idx).border = border
        ws1.cell(row=row, column=2, value=w['display_name'] or 'Noma\'lum').border = border
        ws1.cell(row=row, column=3, value=f"@{w['username']}" if w['username'] else '—').border = border
        ws1.cell(row=row, column=4, value=w['telegram_id']).border = border
        ws1.cell(row=row, column=5, value=w['order_count']).border = border
        ws1.cell(row=row, column=6, value=w['total_earnings']).border = border
        cell = ws1.cell(row=row, column=7, value=-w['commission_amount'])
        cell.border = border
        cell.font = Font(color="ef4444")
        cell = ws1.cell(row=row, column=8, value=w['service_fee_total'])
        cell.border = border
        cell.font = service_font
        cell = ws1.cell(row=row, column=9, value=w['net_payment'])
        cell.border = border
        cell.font = money_font
        current_row += 1

    # Totals row
    total_earnings = sum(w['total_earnings'] for w in workers)
    total_commission = sum(w['commission_amount'] for w in workers)
    total_service_fees = sum(w['service_fee_total'] for w in workers)
    total_net = sum(w['net_payment'] for w in workers)

    ws1.cell(row=current_row, column=1, value="").border = border
    ws1.cell(row=current_row, column=2, value="JAMI").font = Font(bold=True)
    ws1.cell(row=current_row, column=2).border = border
    ws1.cell(row=current_row, column=3, value="").border = border
    ws1.cell(row=current_row, column=4, value="").border = border
    ws1.cell(row=current_row, column=5, value=sum(w['order_count'] for w in workers)).border = border
    ws1.cell(row=current_row, column=5).font = Font(bold=True)
    ws1.cell(row=current_row, column=6, value=total_earnings).border = border
    ws1.cell(row=current_row, column=6).font = Font(bold=True)
    cell = ws1.cell(row=current_row, column=7, value=-total_commission)
    cell.border = border
    cell.font = Font(bold=True, color="ef4444")
    cell = ws1.cell(row=current_row, column=8, value=total_service_fees)
    cell.border = border
    cell.font = Font(bold=True, color="8b5cf6")
    cell = ws1.cell(row=current_row, column=9, value=total_net)
    cell.border = border
    cell.font = Font(bold=True, color="059669", size=12)

    # Adjust column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 15
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 15

    # ===== Sheet 2: Detailed Orders =====
    ws2 = wb.create_sheet(title="Buyurtmalar tafsiloti")

    # Headers
    order_headers = ['#', 'Buyurtma ID', 'Qabriston', 'Marhum ismi', 'Summa', 'Sana', 'Ishchi', 'Username']
    for col, header in enumerate(order_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Orders data
    for idx, o in enumerate(orders, 1):
        row = idx + 1
        ws2.cell(row=row, column=1, value=idx).border = border
        ws2.cell(row=row, column=2, value=o['original_id']).border = border
        ws2.cell(row=row, column=3, value=o['cemetery_name'] or '—').border = border
        ws2.cell(row=row, column=4, value=o['deceased_full_name'] or '—').border = border
        ws2.cell(row=row, column=5, value=o['total_price']).border = border
        ws2.cell(row=row, column=6, value=o['completed_at'].strftime('%d.%m.%Y %H:%M') if o['completed_at'] else '—').border = border
        ws2.cell(row=row, column=7, value=o['worker_name'] or '—').border = border
        ws2.cell(row=row, column=8, value=f"@{o['worker_username']}" if o['worker_username'] else '—').border = border

    # Adjust column widths
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 25
    ws2.column_dimensions['D'].width = 25
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 20
    ws2.column_dimensions['H'].width = 18

    # ===== Sheet 3: Service Fees =====
    if active_service_fees:
        ws3 = wb.create_sheet(title="Hizmat to'lovlari")

        ws3['A1'] = "HIZMAT TO'LOVLARI"
        ws3['A1'].font = title_font

        headers = ['#', 'Hizmat nomi', 'Summa (so\'m)', 'Tavsif']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = service_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for idx, sf in enumerate(active_service_fees, 1):
            row = 3 + idx
            ws3.cell(row=row, column=1, value=idx).border = border
            ws3.cell(row=row, column=2, value=sf.service_name).border = border
            ws3.cell(row=row, column=3, value=float(sf.fee_amount)).border = border
            ws3.cell(row=row, column=4, value=sf.description or '—').border = border

        ws3.column_dimensions['A'].width = 5
        ws3.column_dimensions['B'].width = 25
        ws3.column_dimensions['C'].width = 15
        ws3.column_dimensions['D'].width = 40

    # Create response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"hisobot_{date_from.strftime('%Y%m%d')}_{(date_to - timedelta(days=1)).strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
